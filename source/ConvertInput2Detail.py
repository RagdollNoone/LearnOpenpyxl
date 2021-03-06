#!/usr/bin/env python
# -*- coding:utf-8 -*-

import os
import shutil

from Static import ignoreFileName
from Static import targetNameDict
from Static import source_col
from Static import source_begin_row
from Static import source_end_row
from Static import targetEvaluateValueDict
from Static import template_detail_table_path
from Static import template_overview_table_name
from Static import template_overview_table_path
from Static import targetSheetDict
from Static import template_detail_table_name


class ConvertInput2Detail():
    DataMgr = None

    def __init__(self):
        return

    def set_data_mgr(self, data_mgr):
        self.DataMgr = data_mgr

    def convert(self, file_dir):
        file_list = os.listdir(file_dir)

        for i in range(0, len(file_list)):
            path = os.path.join(file_dir, file_list[i])
            base_name = os.path.basename(path)

            if self.check_is_ignore_file(base_name):
                continue

            if os.path.isdir(path):
                self.convert(path)

            file_format_arr = os.path.splitext(path)
            if file_format_arr[1] == ".xlsx":
                print("convert file name is " + path)

                identity = self.get_file_identity(base_name)
                if identity == "illegal identity":
                    print("Can not recognize the file " + base_name + " identity")
                    return

                can_direct_write = self.check_can_direct_write(identity)
                target_file_path = self.prepare_for_output(path)
                self.data_operate(path, identity, can_direct_write, target_file_path)
        return

    @staticmethod
    def check_evaluate_string_legal(evaluate):
        result = False

        for key in targetEvaluateValueDict:
            if key == evaluate:
                return True

        return result

    @staticmethod
    def get_evaluate_value(evaluate):
        for key in targetEvaluateValueDict:
            if key == evaluate:
                return targetEvaluateValueDict[key]

        return evaluate

    @staticmethod
    def check_is_ignore_file(name):
        for i in range(len(ignoreFileName)):
            if name == ignoreFileName[i]:
                return True

        return False

    @staticmethod
    def check_can_direct_write(name):
        if not (name in targetNameDict):
            return -1

        return targetNameDict[name][1]

    @staticmethod
    def get_file_identity(file_name):
        pre = file_name.split('2')[0]
        if "Boss" in pre:
            return "Boss"

        if "Peer" in pre:
            return "Peer"

        if "DirectReport" in pre:
            return "DirectReport"

        if "Self" in pre:
            return "Self"

        return "illegal identity"

    """
    生成输出目录和文件
    """
    @staticmethod
    def prepare_for_output(path):
        parent_folder_name = os.path.dirname(path)
        target_folder_path = parent_folder_name.replace("input", "output")
        target_file_path = os.path.join(target_folder_path, template_detail_table_name)

        if not os.path.exists(target_folder_path):
            os.makedirs(target_folder_path)

        if not os.path.exists(target_file_path):
            shutil.copyfile(template_detail_table_path, target_file_path)

        overview_file_path = os.path.join(target_folder_path, template_overview_table_name)
        if not os.path.exists(overview_file_path):
            shutil.copyfile(template_overview_table_path, overview_file_path)

        return target_file_path

    def data_operate(self, path, evaluate_identity, can_direct_write, target_file_path):
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active

        for j in range(0, source_end_row - source_begin_row + 1):
            read_unit = source_col + str(j + source_begin_row)

            if self.check_evaluate_string_legal(ws[read_unit].value):
                # print(ws[read_unit].value)
                base_name = os.path.basename(path)
                be_evaluated_name = self.get_be_evaluated_name(base_name)

                if can_direct_write:
                    value = self.write_data(evaluate_identity, target_file_path, ws[read_unit].value, j)
                    self.DataMgr.process_score(be_evaluated_name, evaluate_identity, value)
                else:
                    value = self.get_evaluate_value(ws[read_unit].value)
                    self.DataMgr.process_peer_score(be_evaluated_name, j, value)
            else:
                print("illegal evaluate in file " + os.path.abspath(path) + " unit " + read_unit)
                return

        return

    def write_data(self, evaluate_identity, target_file_path, value, index):
        from openpyxl import load_workbook
        wb_write = load_workbook(target_file_path)
        write_sheet, write_unit = self.get_target_sheet_and_unit(evaluate_identity, index)

        ws_write = wb_write[write_sheet]
        ws_write[write_unit] = self.get_evaluate_value(value)
        wb_write.template = False
        wb_write.save(target_file_path)

        return ws_write[write_unit].value

    @staticmethod
    def get_target_sheet_and_unit(evaluate_identity, index):
        for key in targetSheetDict:
            if index >= targetSheetDict[key][2] and index <= targetSheetDict[key][3]:
                offset = index - targetSheetDict[key][2]
                target_unit = str(targetNameDict[evaluate_identity][0]) + str(targetSheetDict[key][0] + offset)
                return key, target_unit

        return "illegal_sheet", -1

    @staticmethod
    def get_be_evaluated_name(file_name):
        be_evaluate_file_name = file_name.split('2')[1]
        return be_evaluate_file_name.split('xlsx')[0]

    def generate_other_average(self, file_dir):
        file_list = os.listdir(file_dir)

        for i in range(0, len(file_list)):
            path = os.path.join(file_dir, file_list[i])
            base_name = os.path.basename(path)

            if os.path.isdir(path):
                self.generate_other_average(path)

            if base_name == template_detail_table_name:
                print("generate_other_average file name is " + path)

                from openpyxl import load_workbook
                wb = load_workbook(path)
                for key in targetSheetDict:
                    ws = wb[key]
                    for j in range(int(targetSheetDict[key][0]), int(targetSheetDict[key][1] + 1)):
                        # print("generate_other_average current sheet is " + key + "current process row is " + str(j))

                        direct_value = ws[('D' + str(j))].value
                        peer_value = ws[('E' + str(j))].value
                        boss_value = ws[('F' + str(j))].value
                        value = int((direct_value + peer_value + boss_value) / 3)
                        ws[('G' + str(j))] = value
                        wb.template = False
                        wb.save(path)
        return

