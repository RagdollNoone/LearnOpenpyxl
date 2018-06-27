#!/usr/bin/env python
# -*- coding:utf-8 -*-

import os
import shutil


class ReadData():
    rootDir = '../Resource/input'

    targetName = 'target.xlsx'
    targetStringArr = {'Strongly Disagree': 0,
                       'Disagree': 20,
                       'Slightly Disagree': 40,
                       'Slightly Agree': 60,
                       'Agree': 80,
                       'Strongly Agree': 100}

    source_col = 'C'
    source_begin_row = 16
    source_end_row = 30

    targetRootDir = '../Resource/output'
    target_col = 'L'
    target_begin_row = 3
    target_end_row = 17

    template_target_path = '../Resource/target.xlsx'

    def __init__(self):
        return

    def read_file(self, file_dir):
        file_list = os.listdir(file_dir)

        for i in range(0, len(file_list)):
            path = os.path.join(file_dir, file_list[i])

            if os.path.basename(path) == self.targetName:
                continue

            if os.path.isdir(path):
                self.read_file(path)

            file_format_arr = os.path.splitext(path)
            if file_format_arr[1] == ".xlsx":

                from openpyxl import load_workbook
                wb = load_workbook(path)
                ws = wb.active

                for j in range(0, self.source_end_row - self.source_begin_row + 1):
                    read_unit = self.source_col + str(j + self.source_begin_row)

                    if self.check_evaluate_string_legal(ws[read_unit].value):
                        print(ws[read_unit].value)

                        parent_folder_name = os.path.dirname(path)
                        target_folder_path = parent_folder_name.replace("input", "output")
                        target_file_path = os.path.join(target_folder_path, "target.xlsx")

                        if not os.path.exists(target_folder_path):
                            os.makedirs(target_folder_path)

                        if not os.path.exists(target_file_path):
                            shutil.copyfile(self.template_target_path, target_file_path)

                        from openpyxl import load_workbook
                        wb_write = load_workbook(target_file_path)
                        ws_write = wb_write.active

                        write_unit = self.target_col + str(j + self.target_begin_row)

                        ws_write[write_unit] = self.get_evaluate_value(ws[read_unit].value)
                        wb_write.template = False
                        wb_write.save(target_file_path)

                    else:
                        print("illegal evaluate in file " + os.path.abspath(path) + " unit " + read_unit)
                        return

        return

    def check_evaluate_string_legal(self, evaluate):
        result = False

        for key in self.targetStringArr:
            if key == evaluate:
                return True

        return result

    def get_evaluate_value(self, evaluate):
        for key in self.targetStringArr:
            if key == evaluate:
                return self.targetStringArr[key]

        return -1


read_data = ReadData()
read_data.read_file(read_data.rootDir)



