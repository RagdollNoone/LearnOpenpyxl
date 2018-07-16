#!/usr/bin/env python 
# -*- coding:utf-8 -*-

from Base import MISSING
from Base import SPLIT_CHAR_EQUAL
from openpyxl import load_workbook


class BaseXlsx(object):

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = MISSING
        self.current_ws = MISSING
        self.row_name = {}
        self.col_name = {}
        self.top_left_str = MISSING
        self.sheet_name_arr = {}
        self.key_cell_table = {}

    def active_xlsx(self, config_path, sheet_name=MISSING):
        self.wb = load_workbook(self.file_path)

        if sheet_name is MISSING:
            sheet_name = self.wb.sheetnames[0]

        self.current_ws = self.wb[sheet_name]

        self.top_left_str = self.get_top_left_str(config_path)
        assert self.top_left_str is not MISSING, \
            "get top left str fail\n"

        return

    def load_sheet(self, sheet_name=MISSING):
        """" load sheet by name """

        assert self.wb is not MISSING, "current wb is NULL\n"

        result = MISSING
        if sheet_name is MISSING: result = self.wb.active
        else: result = self.wb[sheet_name]
        self.current_ws = result

        return result

    def write_value(self, cell, value):
        """ write value to sheet """

        assert self.wb is not MISSING, "current wb is NULL\n"

        cell.value = value
        self.wb.template = False
        self.wb.save(self.flie_path)

    def get_cell_by_content(self, content, need_cache=False):
        """ find col index by giving a str """

        if content in self.key_cell_table:
            return self.key_cell_table[content]

        assert self.current_ws is not MISSING, "current ws is NULL\n"

        for rows in self.current_ws.rows:
            for cell in rows:
                if cell.value == content:
                    if need_cache:
                        self.key_cell_table[content] = cell
                    return cell

        print("get_cell_by_content: find cell fail about " + content)
        return MISSING

    def get_cell_by_cell_and_offset(self, cell, row_offset=0, col_offset=0):
        """
        :param cell: start position cell
        :param row_offset: row offset value
        :param col_offset: col offset value
        :return: end position cell
        """

        assert self.current_ws is not MISSING, "current ws is NULL\n"

        c = cell.col_idx + col_offset
        r = cell.row + row_offset
        result = self.current_ws.cell(row=r, col=c)
        return result

    def get_cell_cross(self, cell1, cell2):
        """
        :param cell1: cell contain target cell row info
        :param cell2: cell contain target cell col info
        :return: cell with max(row1, row2), max(col1, col2)
        """

        assert self.current_ws is not MISSING, "current ws is NULL\n"

        c1 = cell1.col_idx
        c2 = cell2.col_idx

        r1 = cell1.row
        r2 = cell2.row
        result = self.current_ws.cell(row=max(r1, r2), col=max(c1, c2))
        return result

    def get_sheet_name_arr(self):
        """
        construct sheet dict,
        key is index to help loop find next
        """

        assert self.wb is not MISSING, "current wb is NULL\n"

        i = 0
        for name in self.wb.sheetnames:
            self.sheet_name_arr[i] = name

    def get_next_sheet(self):
        """ loop find next sheet and active it"""

        for key in self.sheet_name_arr:
            if self.current_ws.title == self.sheet_name_arr[key]:
                next_key = (key + 1) % len(self.sheet_name_arr)
                next_sheet_name = self.sheet_name_arr[next_key]
                self.load_sheet(next_sheet_name)
                break

    def get_sheet_range(self):
        """ get table valid area row number """

        assert self.current_ws is not MISSING, \
            "current ws is NULL\n"

        top_left_cell = self.get_cell_by_content(self.top_left_str)
        return self.current_ws.max_row - top_left_cell.row


    @staticmethod
    def get_top_left_str(file_path):
        """
        top left str is a coordinate about
        begin col index and row index
        """

        result = MISSING
        f = open(file_path)
        line = f.readline()
        while line:
            arr = line.split(SPLIT_CHAR_EQUAL)
            if len(arr) > 1:
                result = arr[1].strip('\n');
                f.close()
                return result

        f.close()
        return result


# overview_path = "../Resource/RolesOfLeadershipOverview.xlsx"
# detail_path = "../Resource/RolesOfLeadershipDetail.xlsx"
# detail_config_path = "../Resource/DetailExcelConfig.txt"
#
# detail_table = BaseXlsx(detail_path)
# overview_table = BaseXlsx(overview_path)
#
# detail_table.active_xlsx()
# detail_table.top_left_str = detail_table.get_top_left_str(detail_config_path)
# cell_result1 = detail_table.get_cell_by_content(detail_table.top_left_str, need_cache=True)
# cell_result2 = detail_table.get_cell_by_content(detail_table.top_left_str)

