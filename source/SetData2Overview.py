#!/usr/bin/env python 
# -*- coding:utf-8 -*-

import os

from Static import template_overview_table_name
from Static import template_detail_table_name
from Static import targetSheetArr
from Static import targetSheetDict
from Static import overviewNameDict
from Static import targetNameDict
from Static import radarOverviewNameDict


class SetData2Overview():
    def __init(self):
        return

    def generate_overview_table(self, file_dir):
        file_list = os.listdir(file_dir)

        for i in range(0, len(file_list)):
            path = os.path.join(file_dir, file_list[i])
            base_name = os.path.basename(path)

            if os.path.isdir(path):
                self.generate_overview_table(path)

            if base_name == template_overview_table_name:
                print("generate_overview_table file name is " + path)

                from openpyxl import load_workbook
                to_wb = load_workbook(path)
                from_path = path.replace(template_overview_table_name, template_detail_table_name)
                from_wb = load_workbook(from_path)

                for j in range(len(targetSheetArr)):
                    sheet_name = targetSheetArr[j]
                    from_ws = from_wb[sheet_name]
                    begin_row = targetSheetDict[sheet_name][0]
                    end_row = targetSheetDict[sheet_name][1]

                    peer_total = 0
                    boss_total = 0
                    self_total = 0
                    direct_report_total = 0

                    for k in range(begin_row, end_row):
                        peer_total = peer_total + from_ws[self.get_unit('Peer', k)].value
                        boss_total = boss_total + from_ws[self.get_unit('Boss', k)].value
                        self_total = self_total + from_ws[self.get_unit('Self', k)].value
                        direct_report_total = direct_report_total + from_ws[self.get_unit('DirectReport', k)].value

                    number = end_row - begin_row + 1
                    peer_average = int(peer_total / number)
                    boss_average = int(boss_total / number)
                    self_average = int(self_total / number)
                    direct_report_average = int(direct_report_total / number)

                    to_ws = to_wb.active
                    to_ws[overviewNameDict['Peer'][sheet_name]] = peer_average
                    to_ws[overviewNameDict['Boss'][sheet_name]] = boss_average
                    to_ws[overviewNameDict['Self'][sheet_name]] = self_average
                    to_ws[overviewNameDict['DirectReport'][sheet_name]] = direct_report_average

                    radar_ws = to_wb["Sheet2"]
                    radar_ws[radarOverviewNameDict['Peer'][sheet_name]] = peer_average
                    radar_ws[radarOverviewNameDict['Boss'][sheet_name]] = boss_average
                    radar_ws[radarOverviewNameDict['Self'][sheet_name]] = self_average
                    radar_ws[radarOverviewNameDict['DirectReport'][sheet_name]] = direct_report_average

                # to_wb.template = False
                # to_wb.save(path)
        return

    @staticmethod
    def get_unit(identity, index):
        col = targetNameDict[identity][0]
        result = col + str(index)
        return result

    @staticmethod
    def generate_line_chart(file_dir):
        file_list = os.listdir(file_dir)

        for i in range(0, len(file_list)):
            path = os.path.join(file_dir, file_list[i])
            base_name = os.path.basename(path)

            if os.path.isdir(path):
                SetData2Overview.generate_line_chart(path)

            if base_name == template_overview_table_name:
                print("generate_line_chart file name is " + path)

                from openpyxl.chart import (
                    LineChart,
                    Reference,
                )

                from openpyxl import load_workbook
                wb = load_workbook(path)
                ws = wb.active
                labels = Reference(ws, min_col=5, min_row=10, max_col=8)
                data = Reference(ws, min_col=4, min_row=11, max_col=8, max_row=14)

                # Chart with date axis
                line_chart = LineChart()
                line_chart.style = 26  # 17 黑白
                line_chart.width = 30
                line_chart.height = 14

                line_chart.add_data(data, from_rows=True, titles_from_data=True)
                line_chart.set_categories(labels)

                for s in line_chart.series:
                    s.graphicalProperties.line.width = 20000

                ws.add_chart(line_chart, "D21")

                wb.template = False
                wb.save(path)
        return

    @staticmethod
    def generate_radar_chart(file_dir):
        file_list = os.listdir(file_dir)

        for i in range(0, len(file_list)):
            path = os.path.join(file_dir, file_list[i])
            base_name = os.path.basename(path)

            if os.path.isdir(path):
                SetData2Overview.generate_radar_chart(path)

            if base_name == template_overview_table_name:
                print("generate_radar_chart file name is " + path)

                from openpyxl.chart import (
                    RadarChart,
                    Reference,
                )

                from openpyxl import load_workbook
                wb = load_workbook(path)
                ws = wb["Sheet2"]
                to_ws = wb.active
                labels = Reference(ws, min_col=5, min_row=10, max_col=8)
                data = Reference(ws, min_col=4, min_row=11, max_col=8, max_row=14)

                radar_chart = RadarChart()
                radar_chart.type = "marker"

                radar_chart.add_data(data, from_rows=True, titles_from_data=True)
                radar_chart.set_categories(labels)

                radar_chart.style = 26
                radar_chart.y_axis.delete = True
                radar_chart.width = 30
                radar_chart.height = 14

                for s in radar_chart.series:
                    s.graphicalProperties.line.width = 20000

                to_ws.add_chart(radar_chart, "D58")

                wb.template = False
                wb.save(path)
        return
